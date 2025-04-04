
# ezt a Gemini Code assist csinálta az eredeti orderform-ból, de nem jó xlsx-et exportál sajnos


import pyodbc
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from auDAOlib import readPROD  # Assuming this is a custom library


arlista = 23
akciodatum = r"'2025-03-10'"
nemakciodatum = r"'2025-03-23'"

picPath = 'D:\\Arsuna\\TERMEKKEPEK\\ppp200\\'
xlsPath = 'D:\\Arsuna\\AUrendeles2025.xlsx'
xlsPathout = f'D:\\Arsuna\\AUrendeles2025-{arlista}.xlsx'

# SQL Query (no changes needed here)
sqlQuery = f"""
select top 10 i.itemcode, i.codebars, i.itemname, min(i.salpackun) csomag, 
min(dbo.listaar({arlista},{nemakciodatum}, i.itemcode)) nkar, 
min(dbo.listaar({arlista},{akciodatum}, i.itemcode)) kedvar, 
sum(be.OpenQty) beszerzes, coalesce(min(o.beerkezes), min(be.docduedate), '2024.05.01') bedatum, i.u_id, sum(i.onhand) keszlet, i.ItmsGrpCod
from ARSUNA_2020_PROD.dbo.oitm i 
left join AUassist.dbo.orderform o on o.cikkszam=i.itemcode
left join
    (select pl.itemcode, pl.OpenQty, iif(ph.docduedate>getdate(),ph.DocDueDate,null) docduedate from ARSUNA_2020_PROD.dbo.por1 pl join ARSUNA_2020_PROD.dbo.opor ph on ph.DocEntry=pl.DocEntry and ph.DocStatus='O' where pl.LineStatus='O'
    union all
    select bl.itemcode, bl.OpenQty, iif(bh.vatdate>getdate(),bh.vatdate,null) docduedate from ARSUNA_2020_PROD.dbo.pch1 bl join ARSUNA_2020_PROD.dbo.opch bh on bh.DocEntry=bl.DocEntry and bh.DocStatus='O' where bl.LineStatus='O' and bh.isins='Y') be on be.ItemCode=i.ItemCode
where i.itemname not like '%kulacstet%' and i.itemname not like '%szilikon%' and i.ItmsGrpCod not in (103) and i.itemcode not in ('1004') and u_id not in ('581', '633', '637')
group by i.ItemCode, i.CodeBars, i.ItemName, i.u_id,  i.ItmsGrpCod
having (sum(be.OpenQty)>0 or i.itemcode in (select cikkszam from AUassist.dbo.orderform union select id from auassist.dbo.temp))
and (sum(i.onhand)=0 or i.itemcode in (select id from auassist.dbo.temp))

union all
select ean8, ean13, cikknev, csomegys, 
case {arlista} when 1 then nk when 18 then ot when 20 then tiz when 21 then tizenot when 23 then husz end ar, case {arlista} when 1 then nk when 18 then ot when 20 then tiz when 21 then tizenot when 23 then husz end*.93 kedvar,
0,beerkezes, id, 0, null
from AUassist.dbo.orderform_extra

"""

# Fetch data from the database
dfStock = readPROD(sqlQuery)
dfStock.sort_values(by=['u_id'], inplace=True)

if not isinstance(dfStock, pd.DataFrame):
    print(dfStock)
    sys.exit()

# Load existing workbook or create a new one
try:
    workbook = load_workbook(xlsPath)
except FileNotFoundError:
    workbook = Workbook()

# Select the worksheet
try:
    worksheet = workbook['ArsUna']
except KeyError:
    worksheet = workbook.create_sheet('ArsUna')


# Write the DataFrame to the worksheet
# Optionally, you can add a header to the sheet if it's empty
if worksheet.max_row == 1 and worksheet.max_column == 1:
    headers = list(dfStock.columns)
    worksheet.append(headers)

for r in dataframe_to_rows(dfStock, index=False, header=False):
    worksheet.append(r)


# Add formulas and pictures after the data.
for index, row in dfStock.iterrows():
    rownum = index + 2 + (1 if worksheet.max_row > 1 else 0) # add header offset if needed

    # Add the formula in column 9
    worksheet.cell(row=rownum, column=9).value = f'=RC[-3]*RC[-1]'

    # Add the picture
    imagefile = picPath + row['itemcode'] + '.jpg'
    try:
        img = Image(imagefile)
        # Scale the image if needed
        max_width = 100
        max_height = 100
        width, height = img.width, img.height
        if width > max_width or height > max_height:
          if width/max_width > height/max_height:
              ratio = width/max_width
          else:
              ratio = height/max_height

          img.width = width/ratio
          img.height = height/ratio

        worksheet.add_image(img, f'J{rownum}')  # Place image in column J
    except FileNotFoundError:
        print(f"Image not found: {imagefile}")
    except Exception as e:
        print(f"Error adding image {imagefile}: {e}")


# Save the workbook
workbook.save(xlsPathout)
print(f"Data and formulas written to: {xlsPathout}")