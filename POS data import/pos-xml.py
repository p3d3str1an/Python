from auDAOlib import readPROD
from datetime import datetime, timedelta, date
import argparse, sys
import numpy as np
import pandas as pd
import openpyxl
import os
from credentials import KOZOSPATH
import xml.etree.ElementTree as ET


def myexcepthook(type, value, traceback, oldhook=sys.excepthook):
    oldhook(type, value, traceback)
    input("Press RETURN. ")    # use input() in Python 3.x
sys.excepthook = myexcepthook

exportfolder = KOZOSPATH+r'\PÉNZÜGY\BEOLVASÓK\POS'

## debugold a megfelelő menüponttal (debugikon, legördülőmenü, vagyis custom launch.json, amiben paraméter van a fájlhoz) 

#argumentumnként megadható a fájl, elérési úttal (drag-and-drop is műk)
parser = argparse.ArgumentParser('SBO számlákat párosít POS terminál jelentés-sorokhoz') 
parser.add_argument('poslistname', help='Terminál jelentés fájl', nargs='+')                         
args = parser.parse_args()       

sourcefiles = args.poslistname

#időkülönbség max:
maxtimedelta=timedelta(minutes=5)

for sourcefile in sourcefiles:
	try:
		sourcefolder, filename = os.path.split(sourcefile)
	except:
		input('valami gond a fájllal, kérlek ellenőrizd - lehet hogy nincs is fájl?')
		sys.exit


	# Load and parse the XML file
	tree = ET.parse(sourcefile)
	root = tree.getroot()

	statement_date = root.find('.//statement_date').text

	# A list to hold all the shop records
	all_records = []

	# Find all 'shop_record' elements and extract their data
	for record in root.findall('.//shop_record'):
		record_data = {
			'date': record.find('date').text,
			'time': record.find('time').text,
			'rid': int(record.find('transaction_id').text),
			'osszeg': int(float(record.find('gross_transaction_amount_terminal_currency').text.replace(',', '.')))
	}
		all_records.append(record_data)

	# Create a pandas DataFrame from the list of records
	poslist = pd.DataFrame(all_records)

	#ha a KH-s listát a banki referencia nevén mentjük el, egyből azt teszi a ref-be és a fájlnévbe is
	reference = filename.replace('.xml', '')
	exportfile = exportfolder+'\\'+reference+'-teszt_posexport.xlsx'

	#több napot is átölelhet a KH-s lista, főleg ha szombaton is nyitva voltunk és hétfőn töltjük le
	currdatelist=poslist['date'].unique()

	#manuális indexálás, mert ahány nap, annyi belső for, de az indexnek át kell ölelnie--- ez se túl pythonista
	index=0
	#a mezőfejlécek a DTW miatt duplán szerepelnek
	expheader=[{'DocNum':'DocNum','DocDate':'DocDate', 'CardCode':'CardCode','TransferAccount':'TransferAccount','TransferSum':'TransferSum', 'TransferDate':'TransferDate', 'TransferReference':'TransferReference','Reference2':'Reference2', 'CounterReference':'CounterReference', 'JournalRemarks':'JournalRemarks','TaxDate':'TaxDate', 'DocObjectCode':'DocObjectCode'}]
	explines= [{'parentkey':'docnum', 'linenum':'linenum', 'docentry':'docentry', 'SumApplied':'SumApplied'}]
	fullszamlalist=pd.DataFrame(columns=('docentry', 'docnum', 'doctotal', 'kiallit', 'cardcode'))
	for currdate in currdatelist:
		szamlaquery = fr"""
		select docentry, docnum, doctotal, dbo.IFSZ_DOC_DATETIME_F(CreateDate, createts) kiallit, cardcode from oinv where docdate='{currdate}' and cardcode in ('bolt', 'boltszla', 'partnerkedv') and u_paymet='ÁTUT'
		union all
		SELECT docentry, docnum, doctotal, dbo.IFSZ_DOC_DATETIME_F(updatedate, updatets) kiallit, cardcode FROM oinv WHERE U_elszam='{currdate}' AND cardcode IN ('web', 'webc') AND u_paymet='ÁTUT'
		"""
		szamlak = readPROD(szamlaquery)
		fullszamlalist=pd.concat([fullszamlalist, szamlak], ignore_index=True)

		for row in poslist[poslist['date']==currdate].sort_values('time').itertuples():
			sum=row.osszeg
			#kikeressük a megfelelő számlát
			matchinginvs=szamlak.loc[szamlak['doctotal']==sum] # időellenőrzés egyelőre kivéve, ha kellene -> &(szamlak['kiallit']-row['letrehozas']<=maxtimedelta)
			if  matchinginvs.shape[0]>0:
				matchinginvindex = matchinginvs.index[0]
				szlarow=szamlak.loc[matchinginvindex]
				#kidobjuk a számlalistából a felhasznált számlát, hogy az esetleges hasonló összegeknél ne ugyanazt a docentryt tegyük bele
				szamlak.drop(matchinginvindex, inplace=True)
				document=szlarow['docentry']
				ccode=szlarow['cardcode']
				explines.append({'parentkey':index, 'linenum':'', 'docentry':document, 'SumApplied':sum})
				#az első sorban benne lesz az utalási dátum és referencia, az összes többi dátum és ref mező arra az egy-egy cellára hivatkozik
				if index==0:
					expheader.append({'DocNum':index,'DocDate':statement_date, 'CardCode':ccode,'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':reference,'Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
				else:
					expheader.append({'DocNum':index,'DocDate':'=$B$3', 'CardCode':ccode,'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':'=$G$3','Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
			else:
				explines.append({'parentkey':index, 'linenum':'', 'docentry':'=INDEX(szamlak!A:A,MATCH(E%s,szamlak!B:B))' % str(index+3), 'SumApplied':sum})
				if index==0:
					expheader.append({'DocNum':index,'DocDate':statement_date, 'CardCode':'=INDEX(szamlak!E:E,MATCH(INDEX(lines!C:C,MATCH(A%s,lines!A:A,0)),szamlak!A:A,0))' % str(index+3),'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':reference,'Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
				else:
					expheader.append({'DocNum':index,'DocDate':'=$B$3', 'CardCode':'=INDEX(szamlak!E:E,MATCH(INDEX(lines!C:C,MATCH(A%s,lines!A:A,0)),szamlak!A:A,0))' % str(index+3),'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':'=$G$3','Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
			index+=1 #csúnya antipythonista manuális index inkrementálás

	with pd.ExcelWriter(exportfile) as writer: 
		pd.DataFrame(expheader).to_excel(writer,sheet_name='header',index=False)
		pd.DataFrame(explines).to_excel(writer,sheet_name='lines',index=False)
		fullszamlalist.to_excel(writer, sheet_name='szamlak',index=False)
		poslist.sort_values(by='time').to_excel(writer, sheet_name='poslist',index=False)


	#cellaformátum és színezés (textnek kell lennie a dátumnak, vagy szénné hullik az egész a módosításakor)

	wb = openpyxl.load_workbook(filename=exportfile)
	ws = wb['header']
	ws['B3'].number_format=openpyxl.styles.numbers.FORMAT_TEXT
	ws['B3'].fill = openpyxl.styles.PatternFill(start_color='C4D79B', end_color='C4D79B',fill_type = "solid")
	ws['G3'].fill = openpyxl.styles.PatternFill(start_color='C4D79B', end_color='C4D79B',fill_type = "solid")
	wb.save(filename=exportfile)