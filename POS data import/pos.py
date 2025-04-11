from auDAOlib import readPROD
from datetime import datetime, timedelta, date
import argparse, sys
import numpy as np
import pandas as pd
import openpyxl
import os
from credentials import KOZOSPATH

def myexcepthook(type, value, traceback, oldhook=sys.excepthook):
    oldhook(type, value, traceback)
    input("Press RETURN. ")    # use input() in Python 3.x
sys.excepthook = myexcepthook

exportfolder = KOZOSPATH+r'\PÉNZÜGY\BEOLVASÓK\POS'

## debugold a megfelelő menüponttal (debugikon, legördülőmenü, vagyis custom launch.json, amiben paraméter van a fájlhoz) 

#argumentumnként megadható a fájl, elérési úttal (drag-and-drop is műk)
parser = argparse.ArgumentParser('SBO számlákat párosít POS terminál jelentés-sorokhoz') 
parser.add_argument('poslistname', help='Terminál jelentés fájl', nargs='?')                         
args = parser.parse_args()       

#időkülönbség max:
maxtimedelta=timedelta(minutes=5)

sourcefile = args.poslistname
try:
	sourcefolder, filename = os.path.split(sourcefile)
except:
	input('valami gond a fájllal, kérlek ellenőrizd - lehet hogy nincs is fájl?')
	sys.exit

#ha a KH-s listát a banki referencia nevén mentjük el, egyből azt teszi a ref-be és a fájlnévbe is
reference = filename.replace('.xlsx', '')
exportfile = exportfolder+'\\'+reference+'-posexport.xlsx'

poslistfull = pd.read_excel(sourcefile, usecols='A,D,H,G', names=['tipus','letrehozas','rid', 'osszeg'], engine='openpyxl')
#lezarsor=poslistfull.loc[poslistfull['tipus']=='Lezárás'] ez egyelőre nem kell, de hátha később
poslist = poslistfull.loc[poslistfull['tipus']=='Eladás'].sort_values(by='letrehozas').copy()

#több napot is átölelhet a KH-s lista, főleg ha szombaton is nyitva voltunk és hétfőn töltjük le
currdatelist=np.unique(np.array([datetime.strptime(dt,"%d.%m.%Y %H:%M:%S").strftime("%Y-%m-%d") for dt in poslist.loc[poslist['tipus']=='Eladás']['letrehozas']]))

#kerekítjük az összegeket, mert egyébként floatként olvasódnak be -- noob megoldás, biztos van pyhtonistább
poslist['osszeg'] = poslist['osszeg'].astype('int')
poslist['rid'] = poslist['rid'].astype('int')
poslist.reset_index()

#manuális indexálás, mert ahány nap, annyi belső for, de az indexnek át kell ölelnie--- ez se túl pythonista
index=0
#a mezőfejlécek a DTW miatt duplán szerepelnek
expheader=[{'DocNum':'DocNum','DocDate':'DocDate', 'CardCode':'CardCode','TransferAccount':'TransferAccount','TransferSum':'TransferSum', 'TransferDate':'TransferDate', 'TransferReference':'TransferReference','Reference2':'Reference2', 'CounterReference':'CounterReference', 'JournalRemarks':'JournalRemarks','TaxDate':'TaxDate', 'DocObjectCode':'DocObjectCode'}]
explines= [{'parentkey':'docnum', 'linenum':'linenum', 'docentry':'docentry', 'SumApplied':'SumApplied'}]
fullszamlalist=pd.DataFrame(columns=('docentry', 'docnum', 'doctotal', 'kiallit', 'cardcode'))
for currdate in currdatelist:
	szamlaquery = fr"""
	select docentry, docnum, doctotal, dbo.IFSZ_DOC_DATETIME_F(CreateDate, createts) kiallit, cardcode from oinv where docdate='{currdate}' and cardcode in ('bolt', 'boltszla', 'partnerkedv') and u_paymet='ÁTUT'
	union all
	SELECT docentry, docnum, doctotal, dbo.IFSZ_DOC_DATETIME_F(updatedate, updatets) kiallit, cardcode FROM oinv WHERE U_elszam='{currdate}' AND cardcode IN ('web', 'webc') AND u_paymet='ÁTUT'
	"""
	szamlak = readPROD(szamlaquery)
	fullszamlalist=pd.concat([fullszamlalist, szamlak], ignore_index=True)
	datefilter = (currdate==datetime.strptime(dd,"%d.%m.%Y %H:%M:%S").strftime("%Y-%m-%d") for dd in poslist['letrehozas'])
	for row in poslist.loc[datefilter].itertuples():
		sum=row.osszeg
		#kikeressük a megfelelő számlát
		matchinginvs=szamlak.loc[szamlak['doctotal']==sum] # időellenőrzés egyelőre kivéve, ha kellene -> &(szamlak['kiallit']-row['letrehozas']<=maxtimedelta)
		if  matchinginvs.shape[0]>0:
			matchinginvindex = matchinginvs.index[0]
			szlarow=szamlak.loc[matchinginvindex]
			#kidobjuk a számlalistából a felhasznált számlát, hogy az esetleges hasonló összegeknél ne ugyanazt a docentryt tegyük bele
			szamlak.drop(matchinginvindex, inplace=True)
			document=szlarow['docentry']
			ccode=szlarow['cardcode']
			explines.append({'parentkey':index, 'linenum':'', 'docentry':document, 'SumApplied':sum})
			#az első sorban benne lesz az utalási dátum és referencia, az összes többi dátum és ref mező arra az egy-egy cellára hivatkozik
			if index==0:
				expheader.append({'DocNum':index,'DocDate':currdate, 'CardCode':ccode,'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':reference,'Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
			else:
				expheader.append({'DocNum':index,'DocDate':'=$B$3', 'CardCode':ccode,'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':'=$G$3','Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
		else:
			explines.append({'parentkey':index, 'linenum':'', 'docentry':'=INDEX(szamlak!A:A,MATCH(E%s,szamlak!B:B))' % str(index+3), 'SumApplied':sum})
			if index==0:
				expheader.append({'DocNum':index,'DocDate':currdate, 'CardCode':'=INDEX(szamlak!E:E,MATCH(INDEX(lines!C:C,MATCH(A%s,lines!A:A,0)),szamlak!A:A,0))' % str(index+3),'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':reference,'Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
			else:
				expheader.append({'DocNum':index,'DocDate':'=$B$3', 'CardCode':'=INDEX(szamlak!E:E,MATCH(INDEX(lines!C:C,MATCH(A%s,lines!A:A,0)),szamlak!A:A,0))' % str(index+3),'TransferAccount':3842000,'TransferSum':sum, 'TransferDate':'=$B$3', 'TransferReference':'=$G$3','Reference2':row.rid, 'CounterReference':'=$G$3', 'JournalRemarks':'Bejövő fizetés -POS','TaxDate':'=$B$3', 'DocObjectCode':24})
		index+=1 #csúnya antipythonista manuális index inkrementálás

with pd.ExcelWriter(exportfile) as writer: 
	pd.DataFrame(expheader).to_excel(writer,sheet_name='header',index=False)
	pd.DataFrame(explines).to_excel(writer,sheet_name='lines',index=False)
	fullszamlalist.to_excel(writer, sheet_name='szamlak',index=False)
	poslistfull.sort_values(by='letrehozas').to_excel(writer, sheet_name='poslist',index=False)

#cellaformátum és színezés (textnek kell lennie a dátumnak, vagy szénné hullik az egész a módosításakor)

wb = openpyxl.load_workbook(filename=exportfile)
ws = wb['header']
ws['B3'].number_format=openpyxl.styles.numbers.FORMAT_TEXT
ws['B3'].fill = openpyxl.styles.PatternFill(start_color='C4D79B', end_color='C4D79B',fill_type = "solid")
ws['G3'].fill = openpyxl.styles.PatternFill(start_color='C4D79B', end_color='C4D79B',fill_type = "solid")
wb.save(filename=exportfile)